from __future__ import annotations

import hashlib
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path("/Users/a01-0218-0033/Downloads/消費CRM.xlsx")
OUTPUT = Path("/Users/a01-0218-0033/Documents/New project/outputs/crm_import_ready.xlsx")

CUSTOMER_HEADERS = ["id", "bd", "name", "taxId", "contact", "email", "phone", "paymentTerms", "address", "notes", "updatedAt"]
ORDER_HEADERS = [
    "id",
    "bd",
    "salesType",
    "industry",
    "agency",
    "brandCustomer",
    "quoteDate",
    "quoteAmount",
    "dataFormat",
    "dataCount",
    "deliveryDate",
    "invoiceIssued",
    "notes",
    "updatedAt",
]
INVOICE_HEADERS = [
    "id",
    "bd",
    "paidDate",
    "orderNo",
    "billingSchedule",
    "issueDate",
    "invoiceNo",
    "customerId",
    "company",
    "taxId",
    "subtotal",
    "totalWithTax",
    "product",
    "recipient",
    "recipientAddress",
    "email",
    "phone",
    "dueDate",
    "status",
    "notes",
    "description",
    "updatedAt",
]
PAYMENT_HEADERS = ["id", "bd", "paidDate", "customerId", "invoiceId", "amount", "method", "notes", "updatedAt"]
ACTIVITY_HEADERS = ["id", "bd", "activityDate", "customerId", "type", "content", "nextFollowUp", "updatedAt"]
CONTRACT_HEADERS = [
    "id",
    "bd",
    "category",
    "companyName",
    "department",
    "contactTitle",
    "contactName",
    "phone",
    "primaryContact",
    "customerStatus",
    "contractType",
    "startDate",
    "endDate",
    "contractPrice",
    "annualValue",
    "billingMode",
    "notes",
    "updatedAt",
]
OPTION_HEADERS = ["key", "values", "updatedAt"]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def stable_id(prefix: str, *parts) -> str:
    joined = "|".join(clean(part) for part in parts if clean(part))
    digest = hashlib.sha1(joined.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def number(value) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    if text.startswith("="):
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def date_text(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        text = str(int(value))
        if re.fullmatch(r"\d{8}", text):
            return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
        return clean(value)
    text = clean(value)
    matches = re.findall(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", text)
    if matches and len(matches) == 1:
        year, month, day = matches[0]
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return text


def money_including_tax(subtotal: float) -> int:
    return round(subtotal * 1.05)


def money_excluding_tax(total: float) -> int:
    return round(total / 1.05) if total else 0


def first_non_empty(*values) -> str:
    for value in values:
        text = clean(value)
        if text:
            return text
    return ""


class Converter:
    def __init__(self):
        self.workbook = load_workbook(SOURCE, data_only=False)
        self.customers: dict[str, dict] = {}
        self.orders: list[dict] = []
        self.invoices: list[dict] = []
        self.payments: list[dict] = []
        self.contracts: list[dict] = []

    def customer_id(self, company, tax_id="") -> str:
        company = clean(company)
        tax_id = clean_tax_id(tax_id)
        if not company:
            return ""
        return stable_id("cus", tax_id or company, company)

    def upsert_customer(self, company, tax_id="", contact="", email="", phone="", address="", notes="", bd="") -> str:
        company = clean(company)
        if not company:
            return ""
        customer_id = self.customer_id(company, tax_id)
        current = self.customers.get(customer_id, {})
        merged = {
            "id": customer_id,
            "bd": first_non_empty(current.get("bd"), bd),
            "name": company,
            "taxId": first_non_empty(current.get("taxId"), clean_tax_id(tax_id)),
            "contact": first_non_empty(current.get("contact"), contact),
            "email": first_non_empty(current.get("email"), email),
            "phone": first_non_empty(current.get("phone"), phone),
            "paymentTerms": first_non_empty(current.get("paymentTerms"), ""),
            "address": first_non_empty(current.get("address"), address),
            "notes": first_non_empty(current.get("notes"), notes),
            "updatedAt": today(),
        }
        self.customers[customer_id] = merged
        return customer_id

    def convert_orders(self):
        ws = self.workbook["委刊單(Order)"]
        for row_idx in range(2, ws.max_row + 1):
            brand = clean(ws.cell(row_idx, 5).value)
            quote_amount = number(ws.cell(row_idx, 7).value)
            if not brand and not quote_amount:
                continue
            notes = clean(ws.cell(row_idx, 12).value)
            self.orders.append(
                {
                    "id": stable_id("ord", "委刊單", row_idx, brand, ws.cell(row_idx, 11).value),
                    "bd": clean(ws.cell(row_idx, 1).value),
                    "salesType": clean(ws.cell(row_idx, 2).value),
                    "industry": clean(ws.cell(row_idx, 3).value),
                    "agency": clean(ws.cell(row_idx, 4).value),
                    "brandCustomer": brand,
                    "quoteDate": date_text(ws.cell(row_idx, 6).value),
                    "quoteAmount": quote_amount,
                    "dataFormat": clean(ws.cell(row_idx, 8).value),
                    "dataCount": number(ws.cell(row_idx, 9).value),
                    "deliveryDate": date_text(ws.cell(row_idx, 11).value),
                    "invoiceIssued": truthy_invoice(ws.cell(row_idx, 13).value),
                    "notes": add_source_note(notes, "委刊單(Order)", row_idx),
                    "updatedAt": today(),
                }
            )

    def convert_invoice_sheet(self, sheet_name: str, header_row: int, source_type: str):
        ws = self.workbook[sheet_name]
        for row_idx in range(header_row + 1, ws.max_row + 1):
            company = clean(ws.cell(row_idx, 7).value)
            bd = clean(ws.cell(row_idx, 1).value)
            invoice_no = clean(ws.cell(row_idx, 6).value)
            order_no = clean_order_no(ws.cell(row_idx, 3).value)
            total_with_tax = number(ws.cell(row_idx, 10).value if source_type == "single" else ws.cell(row_idx, 9).value)
            subtotal = number(ws.cell(row_idx, 9).value) if source_type == "single" else money_excluding_tax(total_with_tax)
            if not company and not invoice_no and not order_no and not total_with_tax:
                continue

            tax_id = clean_tax_id(ws.cell(row_idx, 8).value)
            recipient = clean(ws.cell(row_idx, 12 if source_type == "single" else 11).value)
            address = clean(ws.cell(row_idx, 13 if source_type == "single" else 12).value)
            email = clean(ws.cell(row_idx, 14 if source_type == "single" else 13).value)
            phone = clean(ws.cell(row_idx, 15 if source_type == "single" else 14).value)
            notes = clean(ws.cell(row_idx, 16 if source_type == "single" else 15).value)
            description = clean(ws.cell(row_idx, 17 if source_type == "single" else 16).value)
            customer_id = self.upsert_customer(company, tax_id, recipient, email, phone, address, bd=bd)
            invoice_id = stable_id("inv", sheet_name, row_idx, invoice_no, company, order_no)
            paid_date = date_text(ws.cell(row_idx, 2).value)
            amount_with_tax = total_with_tax or money_including_tax(subtotal)
            status = "已收款" if paid_date else "未收款"

            self.invoices.append(
                {
                    "id": invoice_id,
                    "bd": bd,
                    "paidDate": paid_date,
                    "orderNo": order_no,
                    "billingSchedule": clean(ws.cell(row_idx, 4).value),
                    "issueDate": date_text(ws.cell(row_idx, 5).value),
                    "invoiceNo": invoice_no,
                    "customerId": customer_id,
                    "company": company,
                    "taxId": tax_id,
                    "subtotal": subtotal,
                    "totalWithTax": amount_with_tax,
                    "product": clean(ws.cell(row_idx, 11 if source_type == "single" else 10).value),
                    "recipient": recipient,
                    "recipientAddress": address,
                    "email": email,
                    "phone": phone,
                    "dueDate": "",
                    "status": status,
                    "notes": add_source_note(notes, sheet_name, row_idx),
                    "description": description,
                    "updatedAt": today(),
                }
            )

            if paid_date and amount_with_tax:
                self.payments.append(
                    {
                        "id": stable_id("pay", invoice_id, paid_date, amount_with_tax),
                        "bd": bd,
                        "paidDate": paid_date,
                        "customerId": customer_id,
                        "invoiceId": invoice_id,
                        "amount": amount_with_tax,
                        "method": "",
                        "notes": add_source_note("", sheet_name, row_idx),
                        "updatedAt": today(),
                    }
                )

    def convert_contracts(self):
        ws = self.workbook["合約追蹤"]
        for row_idx in range(2, ws.max_row + 1):
            company = clean(ws.cell(row_idx, 2).value)
            if not company:
                continue
            contact = clean(ws.cell(row_idx, 5).value)
            phone = clean(ws.cell(row_idx, 6).value)
            self.upsert_customer(company, contact=contact, phone=phone, notes="由合約追蹤匯入")
            self.contracts.append(
                {
                    "id": stable_id("con", "合約追蹤", row_idx, company, ws.cell(row_idx, 10).value, ws.cell(row_idx, 11).value),
                    "bd": "",
                    "category": clean(ws.cell(row_idx, 1).value),
                    "companyName": company,
                    "department": clean(ws.cell(row_idx, 3).value),
                    "contactTitle": clean(ws.cell(row_idx, 4).value),
                    "contactName": contact,
                    "phone": phone,
                    "primaryContact": clean(ws.cell(row_idx, 7).value),
                    "customerStatus": clean(ws.cell(row_idx, 8).value),
                    "contractType": clean(ws.cell(row_idx, 9).value),
                    "startDate": date_text(ws.cell(row_idx, 10).value),
                    "endDate": date_text(ws.cell(row_idx, 11).value),
                    "contractPrice": number(ws.cell(row_idx, 12).value),
                    "annualValue": number(ws.cell(row_idx, 13).value),
                    "billingMode": clean(ws.cell(row_idx, 14).value),
                    "notes": add_source_note(clean(ws.cell(row_idx, 15).value), "合約追蹤", row_idx),
                    "updatedAt": today(),
                }
            )

    def run(self):
        self.convert_orders()
        self.convert_invoice_sheet("訂閱型發票及收款紀錄", 2, "subscription")
        self.convert_invoice_sheet("單次型發票及收款紀錄", 1, "single")
        self.convert_contracts()
        write_import_workbook(self)


def clean_tax_id(value) -> str:
    text = clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def clean_order_no(value) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def truthy_invoice(value) -> str:
    if value is True:
        return "是"
    text = clean(value).lower()
    if text in {"true", "yes", "y", "1", "是", "已開"}:
        return "是"
    return "否"


def add_source_note(note: str, sheet_name: str, row_idx: int) -> str:
    source = f"來源：{sheet_name} 第 {row_idx} 列"
    return f"{note}\n{source}".strip() if note else source


def today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def write_import_workbook(converter: Converter):
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Customers", CUSTOMER_HEADERS, list(converter.customers.values()))
    add_sheet(wb, "Orders", ORDER_HEADERS, converter.orders)
    add_sheet(wb, "Invoices", INVOICE_HEADERS, converter.invoices)
    add_sheet(wb, "Payments", PAYMENT_HEADERS, converter.payments)
    add_sheet(wb, "Activities", ACTIVITY_HEADERS, [])
    add_sheet(wb, "Contracts", CONTRACT_HEADERS, converter.contracts)
    add_sheet(
        wb,
        "Options",
        OPTION_HEADERS,
        [
            {"key": "paymentTerms", "values": "現結\n月結 30 天\n月結 45 天\n月結 60 天", "updatedAt": today()},
            {"key": "bd", "values": unique_join(row.get("bd") for row in converter.orders), "updatedAt": today()},
            {"key": "salesType", "values": unique_join(row.get("salesType") for row in converter.orders), "updatedAt": today()},
            {"key": "industry", "values": unique_join(row.get("industry") for row in converter.orders), "updatedAt": today()},
            {"key": "invoiceStatus", "values": "未收款\n部分入款\n已收款\n作廢", "updatedAt": today()},
            {"key": "paymentMethod", "values": "匯款\n支票\n現金\n信用卡", "updatedAt": today()},
            {"key": "customerStatus", "values": unique_join(row.get("customerStatus") for row in converter.contracts), "updatedAt": today()},
            {"key": "billingMode", "values": unique_join(row.get("billingMode") for row in converter.contracts), "updatedAt": today()},
        ],
    )
    wb.save(OUTPUT)


def unique_join(values) -> str:
    seen = []
    for value in values:
        text = clean(value)
        if text and text not in seen:
            seen.append(text)
    return "\n".join(seen)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="F6D85F")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(clean(cell.value)) for cell in column_cells[:200])
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)


if __name__ == "__main__":
    Converter().run()
    print(OUTPUT)
